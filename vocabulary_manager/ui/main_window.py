import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from .folder_tree import FolderTree
from .entry_list import EntryList
from .search import SearchWindow
from ttkthemes import ThemedTk

class MainWindow(ThemedTk):
    def __init__(self):
        super().__init__()
        self.set_theme("black")
        self.title("单词积累助手")
        self.geometry("900x600")

        self.db = None
        self.current_folder_id = None

        self._create_menu()
        self._create_toolbar()
        self._create_main_pane()

        # 状态栏
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # 初始打开最近库或新建
        self.after(100, self._init_database)

    def _create_menu(self):
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="新建库", command=self.new_database)
        file_menu.add_command(label="打开库", command=self.open_database)
        file_menu.add_separator()
        file_menu.add_command(label="从Excel导入...", command=self.import_excel)
        file_menu.add_command(label="导出到Excel...", command=self.export_excel)
        file_menu.add_separator()
        file_menu.add_command(label="清除重复单词", command=self.deduplicate)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.on_exit)
        menubar.add_cascade(label="文件", menu=file_menu)

    def _create_toolbar(self):
        toolbar = ttk.Frame(self)
        toolbar.pack(side=tk.TOP, fill=tk.X, padx=2, pady=2)

        ttk.Button(toolbar, text="新建文件夹", command=self.new_folder).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="新建单词", command=self.new_word).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="删除", command=self.delete_selected).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="重命名", command=self.rename_selected).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="移动", command=self.move_selected).pack(side=tk.LEFT, padx=2)

        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)

        ttk.Label(toolbar, text="搜索:").pack(side=tk.LEFT)
        self.search_entry = ttk.Entry(toolbar, width=20)
        self.search_entry.pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="搜索", command=self.do_search).pack(side=tk.LEFT, padx=2)

    def _create_main_pane(self):
        paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        self.folder_tree = FolderTree(paned, None, self.on_folder_select)
        self.entry_list = EntryList(paned, None)
        paned.add(self.folder_tree, weight=1)
        paned.add(self.entry_list, weight=2)

        # 设置互连
        self.folder_tree.set_new_word_callback(self.on_new_word_in_folder)

    # ---------- 数据库管理 ----------
    def _init_database(self):
        from db import Database
        self.db = Database()
        # 尝试打开最近库
        last_path = None
        # 从注册表/文件获取？这里简单从默认位置读取，不行则新建
        config_file = "last_db.txt"
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                last_path = f.read().strip()
        if last_path and os.path.exists(last_path):
            self.open_database_path(last_path)
        else:
            # 无最近库，提示新建或打开
            choice = messagebox.askyesno("选择数据库", "没有找到最近使用的词汇库，是否新建一个？\n选择“是”新建，选择“否”打开已有库。")
            if choice:
                self.new_database()
            else:
                self.open_database()

    def open_database_path(self, path):
        try:
            self.db.connect(path)
            self.folder_tree.db = self.db
            self.entry_list.db = self.db
            self.folder_tree.load_root()
            self.entry_list.load_entries(1)
            self.current_folder_id = 1
            self.status_var.set(f"当前库: {os.path.basename(path)}")
            # 保存最近路径
            with open("last_db.txt", 'w') as f:
                f.write(path)
        except Exception as e:
            messagebox.showerror("错误", f"无法打开数据库: {e}")

    def new_database(self):
        path = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("数据库文件", "*.db")])
        if not path:
            return
        # 创建空文件
        open(path, 'a').close()
        self.open_database_path(path)

    def open_database(self):
        path = filedialog.askopenfilename(filetypes=[("数据库文件", "*.db")])
        if path:
            self.open_database_path(path)

    # ---------- 导航 ----------
    def on_folder_select(self, folder_id):
        self.current_folder_id = folder_id
        self.entry_list.load_entries(folder_id)
        path = '/'.join(self.db.get_folder_path(folder_id))
        self.status_var.set(f"当前路径: {path}")

    def jump_to_folder(self, folder_id):
        self.folder_tree.select_folder(folder_id)
        self.on_folder_select(folder_id)

    def on_new_word_in_folder(self, folder_id):
        """从文件夹树右键添加单词时调用"""
        self.jump_to_folder(folder_id)
        self.new_word()

    # ---------- 工具栏动作 ----------
    def new_folder(self):
        if self.current_folder_id is None:
            return
        new_id = self.db.add_folder(self.current_folder_id)
        self.folder_tree.refresh_children(self.current_folder_id)
        self.folder_tree._start_rename(new_id)

    def new_word(self):
        if self.current_folder_id is None:
            return
        from .dialogs import EditEntryDialog
        EditEntryDialog(self, "新建单词", callback=self.entry_list._add_entry)

    def delete_selected(self):
        focused = self.focus_get()
        if isinstance(focused, tk.Toplevel):
            return
        if focused == self.folder_tree.tree or focused.master == self.folder_tree.tree:
            if self.folder_tree.get_selected_folder_id():
                self.folder_tree._delete_folder()
        elif focused == self.entry_list.tree or focused.master == self.entry_list.tree:
            if self.entry_list.get_selected_entry_ids():
                self.entry_list._delete_entry()

    def rename_selected(self):
        focused = self.focus_get()
        if hasattr(focused, 'master') and focused.master == self.entry_list.tree:
            self.entry_list._rename_entry()
        elif hasattr(focused, 'master') and focused.master == self.folder_tree.tree:
            self.folder_tree._rename_folder()

    def move_selected(self):
        focused = self.focus_get()
        if hasattr(focused, 'master') and focused.master == self.entry_list.tree:
            self.entry_list._move_entry()
        elif hasattr(focused, 'master') and focused.master == self.folder_tree.tree:
            self.folder_tree._move_folder()

    def do_search(self):
        keywords = self.search_entry.get()
        if not keywords.strip():
            return
        if self.current_folder_id is None:
            return
        results = self.db.search_entries(self.current_folder_id, keywords)
        SearchWindow(self, self.db, results)

    def on_exit(self):
        if self.db:
            self.db.close()
        self.destroy()
    
    def import_excel(self):
        """从三列无表头的Excel导入：文件夹名(空=根)、单词、释义"""
        if not self.db:
            messagebox.showwarning("提示", "请先打开一个词汇库")
            return

        path = filedialog.askopenfilename(
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        if not path:
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("打开失败", f"无法读取Excel文件：{e}")
            return

        imported = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row:
                continue
            # 允许行只有两列或三列
            folder_name = str(row[0]).strip() if row[0] is not None and str(row[0]).strip() != '' else None
            word = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            definition = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''

            if not word:  # 没有单词则跳过
                errors.append(f"第{row_idx}行：缺少单词，已跳过")
                continue

            # 确定目标文件夹id
            if folder_name is None:
                folder_id = 1  # 根文件夹
            else:
                try:
                    folder_id = self.db.ensure_folder(1, folder_name)
                except Exception as e:
                    errors.append(f"第{row_idx}行：创建文件夹‘{folder_name}’失败（{e}）")
                    continue

            # 添加词条
            try:
                self.db.add_entry(folder_id, word, definition)
                imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行：添加单词‘{word}’失败（{e}）")

        wb.close()

        # 刷新界面
        if self.current_folder_id:
            self.entry_list.load_entries(self.current_folder_id)
        self.folder_tree._populate_root()
        if self.current_folder_id:
            self.folder_tree.select_folder(self.current_folder_id)

        # 显示结果
        result_msg = f"导入完成！成功导入 {imported} 个单词。"
        if errors:
            result_msg += f"\n\n出现 {len(errors)} 个错误：\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                result_msg += f"\n... 共 {len(errors)} 条错误，仅显示前10条。"
        messagebox.showinfo("导入结果", result_msg)
    
    def deduplicate(self):
        """一键清除所有文件夹内的重复词条（单词+释义完全相同）"""
        if not self.db:
            messagebox.showwarning("提示", "请先打开一个词汇库")
            return

        if not messagebox.askyesno("确认清除重复",
                                "此操作将扫描所有文件夹，删除每个文件夹内单词和释义完全相同的重复词条（仅保留一条）。\n\n确定要继续吗？"):
            return

        try:
            deleted = self.db.deduplicate_entries()
        except Exception as e:
            messagebox.showerror("错误", f"清除重复时出错：{e}")
            return

        # 刷新当前视图
        if self.current_folder_id:
            self.entry_list.load_entries(self.current_folder_id)
        self.folder_tree._populate_root()

        messagebox.showinfo("完成", f"已清除重复单词，共删除 {deleted} 条记录。")

    def export_excel(self):
        """导出所有单词为三列Excel：文件夹名(空=根)、单词、释义（无表头）"""
        if not self.db:
            messagebox.showwarning("提示", "请先打开一个词汇库")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            title="导出为Excel"
        )
        if not path:
            return

        try:
            c = self.db.conn.cursor()
            # 查询：如果直接父文件夹是根节点(id=1)，则第一列为空字符串，否则用文件夹名
            c.execute('''
                SELECT e.word, e.definition,
                    CASE WHEN f.id = 1 THEN '' ELSE f.name END AS folder_name
                FROM entries e
                JOIN folders f ON e.folder_id = f.id
                ORDER BY folder_name, e.word
            ''')
            rows = c.fetchall()
        except Exception as e:
            messagebox.showerror("错误", f"读取数据失败：{e}")
            return

        if not rows:
            messagebox.showinfo("提示", "词库中没有单词，无需导出。")
            return

        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            # 写入数据（无表头，顺序为：文件夹名, 单词, 释义）
            for folder_name, word, definition in rows:
                ws.append([folder_name, word, definition])
            wb.save(path)
        except Exception as e:
            messagebox.showerror("导出失败", f"写入Excel出错：{e}")
            return

        messagebox.showinfo("导出完成", f"成功导出 {len(rows)} 个单词到：\n{path}")